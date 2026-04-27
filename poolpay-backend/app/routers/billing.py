import io
from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import StreamingResponse
from sqlmodel import select, Session
from datetime import date
from app.db import get_session
from app.models import Client, Invoice, Payment
from app.schemas import BillingGenerate
from app.services.mercadopago_service import MercadoPagoService
from app.services.billing import generate_invoices as generate_invoices_service

router = APIRouter(prefix="/billing", tags=["billing"])

@router.post("/generate")
def generate_invoices(payload: BillingGenerate, session: Session = Depends(get_session)):
    """Generar facturas automáticamente para todos los clientes activos del período"""
    try:
        year, month = map(int, payload.period.split("-"))
        issue_date = date(year, month, 1)

        # Calcular día de vencimiento (evitar días inválidos)
        max_day = 28 if month == 2 else 30
        due_day = min(payload.due_day, max_day)
        due_date = date(year, month, due_day)
    except ValueError:
        raise HTTPException(400, "Formato de período inválido. Usa: YYYY-MM")

    # Obtener clientes activos
    active_clients = session.exec(select(Client).where(Client.is_active == True)).all()

    created = 0
    skipped = 0

    for client in active_clients:
        # Verificar si ya existe factura para este período
        existing = session.exec(
            select(Invoice).where(
                Invoice.client_id == client.id,
                Invoice.period == payload.period
            )
        ).first()

        if existing:
            skipped += 1
            continue

        # Crear factura
        invoice = Invoice(
            client_id=client.id,
            period=payload.period,
            issue_date=issue_date,
            due_date=due_date,
            subtotal=client.price,
            extras=0.0,
            total=client.price,
            status="pendiente"
        )
        session.add(invoice)
        created += 1

    session.commit()

    return {
        "period": payload.period,
        "created": created,
        "skipped": skipped,
        "total_clients": len(active_clients)
    }

@router.get("/summary/{period}")
def get_billing_summary(period: str, session: Session = Depends(get_session)):
    """Obtener resumen de facturación de un período"""
    invoices = session.exec(select(Invoice).where(Invoice.period == period)).all()

    if not invoices:
        return {
            "period": period,
            "total_invoices": 0,
            "total_amount": 0,
            "paid": 0,
            "pending": 0,
            "partial": 0,
            "overdue": 0,
            "collected": 0
        }

    total_amount = sum(inv.total for inv in invoices)
    by_status = {"pendiente": 0, "pagado": 0, "parcial": 0, "vencido": 0}

    for inv in invoices:
        by_status[inv.status] = by_status.get(inv.status, 0) + 1

    # Calcular total cobrado
    invoice_ids = [inv.id for inv in invoices]
    payments = session.exec(
        select(Payment).where(Payment.invoice_id.in_(invoice_ids))
    ).all()
    collected = sum(p.amount for p in payments)

    return {
        "period": period,
        "total_invoices": len(invoices),
        "total_amount": total_amount,
        "paid": by_status["pagado"],
        "pending": by_status["pendiente"],
        "partial": by_status["parcial"],
        "overdue": by_status["vencido"],
        "collected": collected,
        "pending_amount": total_amount - collected
    }

@router.get("/overdue")
def get_overdue_invoices(session: Session = Depends(get_session)):
    """Obtener facturas vencidas (fecha de vencimiento pasada y no pagadas)"""
    today = date.today()

    overdue = session.exec(
        select(Invoice).where(
            Invoice.due_date < today,
            Invoice.status != "pagado"
        ).order_by(Invoice.due_date)
    ).all()

    # Actualizar estado a vencido si es necesario
    for invoice in overdue:
        if invoice.status != "vencido":
            invoice.status = "vencido"
            session.add(invoice)

    if overdue:
        session.commit()

    return overdue

@router.get("/stats")
def get_general_stats(session: Session = Depends(get_session)):
    """Obtener estadísticas generales del sistema"""
    total_clients = session.exec(select(Client)).all()
    active_clients = [c for c in total_clients if c.is_active]

    all_invoices = session.exec(select(Invoice)).all()
    all_payments = session.exec(select(Payment)).all()

    total_billed = sum(inv.total for inv in all_invoices)
    total_collected = sum(p.amount for p in all_payments)

    return {
        "total_clients": len(total_clients),
        "active_clients": len(active_clients),
        "inactive_clients": len(total_clients) - len(active_clients),
        "total_invoices": len(all_invoices),
        "total_payments": len(all_payments),
        "total_billed": total_billed,
        "total_collected": total_collected,
        "pending_collection": total_billed - total_collected
    }


@router.post('/create-payment-links')
def create_payment_links(period: str | None = None, session: Session = Depends(get_session)):
    """Crear links de MercadoPago para facturas pendientes.

    Si se provee `period` (YYYY-MM) solo se procesan las facturas de ese periodo,
    si no, se procesan todas las facturas con estado 'pendiente'.

    Este endpoint NO guarda las preferencias en la base de datos; devuelve
    una lista con el resultado de la creación por factura.
    """
    # Obtener facturas pendientes
    q = select(Invoice).where(Invoice.status == 'pendiente')
    if period:
        q = q.where(Invoice.period == period)

    invoices = session.exec(q).all()
    results = []

    for inv in invoices:
        client = session.get(Client, inv.client_id)
        if not client:
            results.append({"invoice_id": inv.id, "ok": False, "error": "Cliente no encontrado"})
            continue

        # Construir email/contacto
        client_email = client.whatsapp or client.phone or None
        if client_email and "@" not in client_email:
            client_email = f"{client_email}@poolpay.com"
        if not client_email:
            client_email = f"cliente{client.id}@poolpay.local"

        # Llamar a MercadoPago
        mp_res = MercadoPagoService.create_payment_link(
            title=f"Factura #{inv.id} - Período {inv.period}",
            amount=inv.total,
            client_email=client_email,
            external_reference=f"invoice_{inv.id}",
            description=f"Factura automática para {client.name}"
        )

        if mp_res.get('success'):
            results.append({
                "invoice_id": inv.id,
                "ok": True,
                "preference_id": mp_res.get('preference_id'),
                "payment_link": mp_res.get('init_point')
            })
        else:
            results.append({"invoice_id": inv.id, "ok": False, "error": mp_res.get('error')})

    return {"count": len(results), "results": results}


@router.post('/auto-generate-and-create-links')
def auto_generate_and_create_links(payload: BillingGenerate, session: Session = Depends(get_session)):
    """Generar facturas para un periodo y crear links de MercadoPago para las nuevas facturas.

    Retorna cuántas facturas se crearon y los links generados para las nuevas facturas.
    """
    # Generar facturas (usa el servicio que ya existe)
    created = generate_invoices_service(session, payload.period, payload.due_day)

    # Buscar las facturas del periodo que estén pendientes
    q = select(Invoice).where(Invoice.period == payload.period, Invoice.status == 'pendiente')
    invoices = session.exec(q).all()

    # Crear links para las facturas pendientes (incluye las nuevas)
    links_res = []
    for inv in invoices:
        client = session.get(Client, inv.client_id)
        client_email = client.whatsapp or client.phone or None
        if client_email and "@" not in client_email:
            client_email = f"{client_email}@poolpay.com"
        if not client_email:
            client_email = f"cliente{client.id}@poolpay.local"

        mp_res = MercadoPagoService.create_payment_link(
            title=f"Factura #{inv.id} - Período {inv.period}",
            amount=inv.total,
            client_email=client_email,
            external_reference=f"invoice_{inv.id}",
            description=f"Factura generada automáticamente"
        )

        links_res.append({"invoice_id": inv.id, "mercadopago": mp_res})

    return {"created_invoices": created, "links": links_res}


@router.get("/export/excel/{period}")
def export_excel(period: str, session: Session = Depends(get_session)):
    """Descargar resumen mensual de facturación en formato Excel (.xlsx).

    Incluye una hoja con el detalle de facturas y otra con el resumen financiero.
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise HTTPException(500, "openpyxl no está instalado")

    # ── Datos ────────────────────────────────────────────────────────────────
    invoices = session.exec(select(Invoice).where(Invoice.period == period).order_by(Invoice.id)).all()
    if not invoices:
        raise HTTPException(404, f"No hay facturas para el período {period}")

    invoice_ids = [inv.id for inv in invoices]
    payments = session.exec(select(Payment).where(Payment.invoice_id.in_(invoice_ids))).all()
    clients = {c.id: c for c in session.exec(select(Client)).all()}

    paid_by_invoice: dict[int, float] = {}
    for p in payments:
        paid_by_invoice[p.invoice_id] = paid_by_invoice.get(p.invoice_id, 0) + p.amount

    # ── Workbook ─────────────────────────────────────────────────────────────
    wb = openpyxl.Workbook()

    # Estilos
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill("solid", fgColor="1E3A5F")
    alt_fill    = PatternFill("solid", fgColor="F0F4F8")
    paid_fill   = PatternFill("solid", fgColor="D1FAE5")
    overdue_fill = PatternFill("solid", fgColor="FEE2E2")
    center = Alignment(horizontal="center", vertical="center")
    thin = Border(
        left=Side(style="thin", color="D1D5DB"),
        right=Side(style="thin", color="D1D5DB"),
        top=Side(style="thin", color="D1D5DB"),
        bottom=Side(style="thin", color="D1D5DB"),
    )

    def style_header_row(ws, row, cols):
        for col in range(1, cols + 1):
            cell = ws.cell(row=row, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center
            cell.border = thin

    def style_data_cell(cell, alt=False, row_fill=None):
        cell.border = thin
        cell.alignment = Alignment(vertical="center")
        if row_fill:
            cell.fill = row_fill
        elif alt:
            cell.fill = alt_fill

    # ── Hoja 1: Detalle de facturas ──────────────────────────────────────────
    ws1 = wb.active
    ws1.title = f"Facturas {period}"
    ws1.row_dimensions[1].height = 14
    ws1.row_dimensions[2].height = 28

    # Título
    ws1.merge_cells("A1:H1")
    title_cell = ws1["A1"]
    title_cell.value = f"RESUMEN DE FACTURACIÓN — {period}"
    title_cell.font = Font(bold=True, size=13, color="1E3A5F")
    title_cell.alignment = center

    headers = ["#", "Cliente", "Barrio", "Plan", "Total", "Pagado", "Pendiente", "Estado"]
    for col, h in enumerate(headers, 1):
        ws1.cell(row=2, column=col, value=h)
    style_header_row(ws1, 2, len(headers))

    for i, inv in enumerate(invoices, 1):
        client = clients.get(inv.client_id)
        paid = paid_by_invoice.get(inv.id, 0.0)
        pending = max(0.0, inv.total - paid)
        row = i + 2
        alt = (i % 2 == 0)

        row_fill = paid_fill if inv.status == "pagado" else (overdue_fill if inv.status == "vencido" else None)

        data = [
            inv.id,
            client.name if client else f"ID {inv.client_id}",
            client.neighborhood if client else "",
            client.plan if client else "",
            inv.total,
            paid,
            pending,
            inv.status.upper(),
        ]
        for col, val in enumerate(data, 1):
            cell = ws1.cell(row=row, column=col, value=val)
            style_data_cell(cell, alt=alt, row_fill=row_fill)
            if col in (5, 6, 7):
                cell.number_format = '"$"#,##0.00'
            if col == 8:
                cell.alignment = Alignment(horizontal="center", vertical="center")

    # Anchos de columna
    col_widths = [6, 30, 20, 12, 14, 14, 14, 12]
    for i, w in enumerate(col_widths, 1):
        ws1.column_dimensions[get_column_letter(i)].width = w

    # ── Hoja 2: Resumen financiero ───────────────────────────────────────────
    ws2 = wb.create_sheet("Resumen")

    total_billed   = sum(inv.total for inv in invoices)
    total_paid     = sum(paid_by_invoice.get(inv.id, 0) for inv in invoices)
    total_pending  = total_billed - total_paid
    count_paid     = sum(1 for inv in invoices if inv.status == "pagado")
    count_pending  = sum(1 for inv in invoices if inv.status == "pendiente")
    count_partial  = sum(1 for inv in invoices if inv.status == "parcial")
    count_overdue  = sum(1 for inv in invoices if inv.status == "vencido")

    ws2.merge_cells("A1:B1")
    ws2["A1"].value = f"RESUMEN — {period}"
    ws2["A1"].font = Font(bold=True, size=13, color="1E3A5F")
    ws2["A1"].alignment = center

    summary_rows = [
        ("Total facturas", len(invoices)),
        ("Total facturado", f"${total_billed:,.2f}"),
        ("Total cobrado",   f"${total_paid:,.2f}"),
        ("Total pendiente", f"${total_pending:,.2f}"),
        ("", ""),
        ("Facturas pagadas",   count_paid),
        ("Facturas pendientes", count_pending),
        ("Facturas parciales",  count_partial),
        ("Facturas vencidas",   count_overdue),
        ("", ""),
        ("Tasa de cobro", f"{round(total_paid/total_billed*100, 1) if total_billed else 0}%"),
    ]

    for i, (label, value) in enumerate(summary_rows, 2):
        ws2.cell(row=i, column=1, value=label).font = Font(bold=bool(label))
        ws2.cell(row=i, column=2, value=value).alignment = Alignment(horizontal="right")

    ws2.column_dimensions["A"].width = 25
    ws2.column_dimensions["B"].width = 18

    # ── Respuesta ────────────────────────────────────────────────────────────
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"poolpay_{period}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
